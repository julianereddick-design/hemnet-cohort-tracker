#!/usr/bin/env python3
"""adcost-report.py — Phase 28 ad-cost reporting (rerunnable).

Pulls the full AdCostV2 history from the shared defaultdb and produces:
  1. <out>/adcost-all-data.xlsx  — every snapshot x muni x tier x price point
     (muni-level detail, color-scaled) so price change is visible at a glance.
  2. <out>/adcost-heatmap.html   — 8-county x {Bas,Plus,Premium,Max} heat map of
     % change vs the previous complete snapshot and vs the fixed anchor, plus the
     weighted ARPL block.
  3. --json                      — the machine-readable shape adcost-report.js
     renders the monthly Slack post from. Printed to STDOUT; every human-readable
     diagnostic goes to stderr so stdout stays parseable.

Pricing basis: PAY_WHEN_LISTING_IS_REMOVED (matches the historical series).
Amounts are as returned by Hemnet's webPricingCalculator, i.e. NET of the 25%
Swedish moms. Percentage changes are VAT-agnostic (the ratio cancels).
County rollup + ARPL weights use the v6 listing mix in data/arpl-baseline.json
(county x tier x price-band listing counts). See docs/ad-cost-scrape-gap.md for the
2026-03-16 -> 2026-06-30 no-backfill gap.

  python scripts/adcost-report.py                       # artifacts into exports/
  python scripts/adcost-report.py --json                # JSON on stdout, no files
  python scripts/adcost-report.py --json --out-dir DIR  # both, one DB pull
"""
import argparse
import datetime
import json
import os
import re
import statistics
import sys

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
import psycopg

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, "exports")
BASELINE = os.path.join(REPO, "data", "arpl-baseline.json")

CORE_TIERS = ["BASIC", "PLUS", "PREMIUM", "MAX"]
# The upsells. NOT packages: they are bought ON TOP of a BASIC..MAX listing, so
# they never belong in the same table as the core fees - summing the two would
# double-count a seller who buys both. Split onto their own sheets 2026-08-18.
ADDON_TIERS = ["PAID_REPUBLISH", "TOPLISTING", "TOPLISTING_5_DAYS"]
ALL_TIERS = CORE_TIERS + ADDON_TIERS

# The ONE anchor every published comparison is made against (decision locked
# 2026-08-17). It is pinned rather than derived as "the last snapshot of 2025"
# so that the Slack post and the heat map it links to can never drift apart, and
# so a future backfill cannot silently move the baseline under a published
# number. 2025-12-21 is verified complete: 420/420 cells, 10 munis.
ANCHOR_DATE = datetime.date(2025, 12, 21)

# Workbook comparison columns start here (Julian, 2026-08-18). The 2025 weekly
# run is still ON the Prices sheet as raw history - what is dropped is the
# per-column CHANGE series, which for 2025 is 30-odd near-identical columns that
# bury the two periods anyone reads. The ANCHOR is deliberately NOT filtered:
# it is the baseline the 'vs Anchor' sheet is defined against, not a column.
WORKBOOK_FROM = datetime.date(2026, 1, 1)
MOMS = 1.25  # Swedish VAT. webPricingCalculator amounts are NET (ex-VAT); the v6
             # Output reports GROSS (inc-moms). net × MOMS ≈ v6 reported figures.

# The grid is defined once, in scripts/lib/adcost_grid.py, and shared with the
# crawler. MUNI is re-shaped to (name, county) here because every existing use
# in this file unpacks a 2-tuple.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib"))
import adcost_grid as _grid

MUNI = {mid: (name, county) for mid, (name, _full, county) in _grid.MUNI.items()}
PRICE_POINTS = list(_grid.PRICE_POINTS)
COUNTIES = list(_grid.COUNTIES)

# A complete run is every muni x every price point x every product.
EXPECTED_CELLS = _grid.EXPECTED_CELLS   # 10 * 6 * 7 = 420

# If the previous complete snapshot is older than this, the period-on-period panel
# renders "n/a (gap)" instead of a number. Sized for the MONTHLY cadence the scrape
# moved to on 2026-08-17 (was 12 days, which made the panel permanently n/a): a
# 31-day month plus slack, but short enough that comparing across the
# 2026-03-16 -> 2026-06-30 outage still refuses to render.
PRIOR_MAX_GAP_DAYS = 40


def load_env():
    env = {}
    for line in open(os.path.join(REPO, ".env"), encoding="utf-8"):
        line = line.strip()
        if line.startswith("export "):
            line = line[7:]
        m = re.match(r"([A-Z_]+)=(.*)", line)
        if m:
            env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
    return env


def fetch_rows(env):
    conn = psycopg.connect(host=env["DB_HOST"], port=env.get("DB_PORT", 5432),
                           user=env["DB_USER"], password=env["DB_PASSWORD"],
                           dbname=env["DB_NAME"], sslmode="require", connect_timeout=15)
    cur = conn.cursor()
    # DEDUPED at the source. 2025-10-19 is a DOUBLE-RUN: 742 raw rows, two per cell.
    # Verified 2026-08-17 that no duplicated cell disagrees on ad_price (0 groups
    # with min <> max), so max() collapses them losslessly. Without this, any
    # "rows == 420" completeness test is wrong and the cube's last-writer-wins
    # depends on row order.
    cur.execute("""select property_municipality_id, property_price, ad_type,
                          max(ad_price) as ad_price,
                          (crawled at time zone 'UTC')::date as d
                   from hemnet_adcostv2
                   where property_municipality_id = any(%s)
                     and property_price = any(%s)
                     and ad_type = any(%s)
                   group by 1, 2, 3, 5
                   order by 5""", (list(MUNI), PRICE_POINTS, ALL_TIERS))
    rows = cur.fetchall()
    conn.close()
    return rows


def build_cube(rows):
    """-> data[date][muni_id][tier][price] = ad_price ; and sorted snapshot dates."""
    data = {}
    for muni_id, price, tier, ad_price, d in rows:
        data.setdefault(d, {}).setdefault(muni_id, {}).setdefault(tier, {})[price] = ad_price
    return data, sorted(data)


# ---------------------------------------------------------------------------
# Snapshot completeness — the guard against phantom price changes
# ---------------------------------------------------------------------------
def cell_count(data_date):
    """Number of (muni, tier, price) observations present on one snapshot date."""
    return sum(len(prices) for tiers in data_date.values() for prices in tiers.values())


def snapshot_stats(data, d):
    cells = cell_count(data[d])
    return {
        "date": d.isoformat(),
        "cells": cells,
        "expected_cells": EXPECTED_CELLS,
        "munis": len(data[d]),
        "complete": cells == EXPECTED_CELLS,
    }


def is_complete(data, d):
    return cell_count(data[d]) == EXPECTED_CELLS


def prior_complete(data, dates, before):
    """Most recent COMPLETE snapshot strictly before `before`, or None.

    Completeness is mandatory, not cosmetic. The 2026-08-02 (42 cells, Stockholm
    only) and 2026-08-09 (35 cells) runs died after the first municipality; taking
    either as a baseline would report ~385 cells as "changed" when they were merely
    never scraped. Missing is not moved.
    """
    for d in reversed([x for x in dates if x < before]):
        if is_complete(data, d):
            return d
    return None


def county_price(data_date, county, tier, price):
    """Average the ad_price across the munis of `county` for (tier, price) on a date."""
    vals = [data_date[mid][tier][price]
            for mid, (_, c) in MUNI.items()
            if c == county and mid in data_date
            and tier in data_date[mid] and price in data_date[mid][tier]]
    return sum(vals) / len(vals) if vals else None


def weighted_price(data_date, baseline, county, tier):
    """Baseline-weighted price for (county, tier) over available price-bands."""
    num = den = 0.0
    bands = baseline.get(county, {}).get(tier, {})
    for band_str, cnt in bands.items():
        p = county_price(data_date, county, tier, int(band_str))
        if p is not None:
            num += cnt * p
            den += cnt
    return (num / den) if den else None


def arpl(data_date, baseline, tiers):
    """Weighted ARPL over all counties x given tiers x bands. -> (blended, per_tier)."""
    per_tier = {}
    tnum = tden = 0.0
    for tier in tiers:
        num = den = 0.0
        for county in COUNTIES:
            for band_str, cnt in baseline.get(county, {}).get(tier, {}).items():
                p = county_price(data_date, county, tier, int(band_str))
                if p is not None:
                    num += cnt * p
                    den += cnt
        per_tier[tier] = (num / den) if den else None
        tnum += num
        tden += den
    return (tnum / tden if tden else None), per_tier


def pct(latest, ref):
    if latest is None or ref is None or ref == 0:
        return None
    return latest / ref - 1.0


# ---------------------------------------------------------------------------
# JSON: what the monthly Slack post is rendered from (adcost-report.js)
# ---------------------------------------------------------------------------
def flatten(data_date):
    """-> {(muni_id, tier, price): ad_price} for one snapshot."""
    return {(mid, tier, price): v
            for mid, tiers in data_date.items()
            for tier, prices in tiers.items()
            for price, v in prices.items()}


def diff_cells(data, latest, anchor):
    """Cells present in BOTH snapshots whose price differs. -> list of dicts.

    Only the intersection is compared, deliberately. A cell absent from either
    side was not observed, and an unobserved cell is not a price change — that
    distinction is the entire defence against a partial run reporting hundreds
    of phantom moves.
    """
    a, l = flatten(data[anchor]), flatten(data[latest])
    out = []
    for key in sorted(a.keys() & l.keys(), key=lambda k: (k[1], k[0], k[2])):
        mid, tier, price = key
        av, lv = float(a[key]), float(l[key])
        if av == lv:
            continue
        out.append({
            "municipality": MUNI[mid][0], "county": MUNI[mid][1], "municipality_id": mid,
            "product": tier, "price_point": price,
            "from": av, "to": lv, "pct": pct(lv, av) * 100.0,
        })
    return out, len(a.keys() & l.keys())


# ---------------------------------------------------------------------------
# County x package matrices — the two tables in the Slack post
# ---------------------------------------------------------------------------
QUARTER_DAYS = 90
# How far the chosen quarter reference may sit from the 90-day target before the
# post stops calling it a quarter and states the real elapsed time instead.
QUARTER_TOLERANCE_DAYS = 30


def pick_ref_near(data, dates, latest, days_back):
    """The COMPLETE snapshot closest to `days_back` before `latest`.

    Returns (date, meta) or (None, meta). Deliberately nearest-to-target rather
    than first-older-than-target: the 2026-03-16 -> 2026-07-01 outage means the
    true 90-day mark can land inside a hole, and the honest answer is "here is
    the closest real snapshot, and here is how far off it actually is" — never a
    fabricated quarter. As monthly snapshots accrue this converges on a real
    quarter by itself (from 2026-12 the 90-day target is 2026-09-01, which exists).
    """
    target = latest - datetime.timedelta(days=days_back)
    candidates = [d for d in dates if d < latest and is_complete(data, d)]
    if not candidates:
        return None, {"target": target.isoformat(), "reason": "no earlier complete snapshot"}
    best = min(candidates, key=lambda d: abs((d - target).days))
    return best, {
        "target": target.isoformat(),
        "actual_days_back": (latest - best).days,
        "off_target_days": abs((best - target).days),
        "on_target": abs((best - target).days) <= QUARTER_TOLERANCE_DAYS,
    }


def county_matrix(data, latest, ref):
    """County x {BASIC,PLUS,PREMIUM,MAX} % change, with Total column and TOTAL row.

    Each cell is the % change in an EQUAL-WEIGHTED BASKET: the mean ad_price over
    that county's municipalities x all six price points. Equal-weighted on purpose
    — the v6 listing-mix weights were dropped as a frozen, drifted one-off, so this
    deliberately carries no weighting a reader would have to trust. It is a price
    index, not a revenue estimate: counties are not scaled by market size.

    Only cells present in BOTH snapshots contribute, so a partial run shrinks the
    basket rather than inventing movement.
    """
    a, l = flatten(data[ref]), flatten(data[latest])
    common = a.keys() & l.keys()

    def pct_for(keys):
        keys = list(keys)
        if not keys:
            return None
        av = sum(float(a[k]) for k in keys) / len(keys)
        lv = sum(float(l[k]) for k in keys) / len(keys)
        p = pct(lv, av)
        return None if p is None else p * 100.0

    rows = []
    for county in COUNTIES:
        in_county = [k for k in common if MUNI[k[0]][1] == county]
        rows.append({
            "county": county,
            "tiers": {t: pct_for(k for k in in_county if k[1] == t) for t in CORE_TIERS},
            "total": pct_for(k for k in in_county if k[1] in CORE_TIERS),
            "cells": len([k for k in in_county if k[1] in CORE_TIERS]),
        })
    return {
        "rows": rows,
        "total_row": {t: pct_for(k for k in common if k[1] == t) for t in CORE_TIERS},
        "grand_total": pct_for(k for k in common if k[1] in CORE_TIERS),
        "cells_compared": len([k for k in common if k[1] in CORE_TIERS]),
        # The three add-ons are not packages and are not in the workbook's table, but
        # they do move (TOPLISTING is up ~18% on the anchor) and would otherwise be
        # invisible. National basket only, computed the SAME way as the matrix cells —
        # mixing a basket figure and a median-of-percentages under one product name
        # puts two different numbers for "BASIC" in one post.
        "addons": {t: pct_for(k for k in common if k[1] == t)
                   for t in ALL_TIERS if t not in CORE_TIERS},
    }


def build_matrices(data, dates, latest, anchor):
    """The two tables: vs the fixed anchor, and vs roughly a quarter back."""
    out = {"anchor": {"ref": anchor.isoformat(),
                      "days_back": (latest - anchor).days,
                      **county_matrix(data, latest, anchor)}}
    qref, qmeta = pick_ref_near(data, dates, latest, QUARTER_DAYS)
    out["quarter"] = ({"ref": qref.isoformat(), "days_back": (latest - qref).days,
                       **qmeta, **county_matrix(data, latest, qref)}
                      if qref else {"ref": None, **qmeta})
    return out


def product_summary(moved, data, latest, anchor):
    """Per-product roll-up: how many of its cells moved, and by how much.

    With a FIXED anchor, "which cells moved" saturates — 420 of 420 have moved
    since 2025-12-21 — so the per-product shape, not the per-cell list, is what
    stays readable month after month. The per-cell list is still emitted in full
    for the months where it is short enough to print.
    """
    a, l = flatten(data[anchor]), flatten(data[latest])
    common = a.keys() & l.keys()
    out = []
    for tier in ALL_TIERS:
        in_tier = [k for k in common if k[1] == tier]
        mv = [m for m in moved if m["product"] == tier]
        pcts = sorted(m["pct"] for m in mv)
        out.append({
            "product": tier,
            "core": tier in CORE_TIERS,
            "compared": len(in_tier),
            "moved": len(mv),
            "median_pct": statistics.median(pcts) if pcts else None,
            "min_pct": pcts[0] if pcts else None,
            "max_pct": pcts[-1] if pcts else None,
            # Municipalities whose price for this product moved — "unchanged in
            # all 10 municipalities" has to be a counted claim, not an assumption.
            "munis_moved": len({m["municipality"] for m in mv}),
            "munis_compared": len({k[0] for k in in_tier}),
        })
    return out


def build_json(data, dates, latest, anchor, report_date):
    """The full machine-readable report. Pure over the cube — no DB, no files."""
    warnings = []
    latest_stats = snapshot_stats(data, latest)
    anchor_stats = snapshot_stats(data, anchor)

    if not latest_stats["complete"]:
        warnings.append(
            f"latest snapshot {latest} is PARTIAL: {latest_stats['cells']}/{EXPECTED_CELLS} cells "
            f"across {latest_stats['munis']} of {len(MUNI)} municipalities — only the cells it does "
            f"carry are compared")
    if not anchor_stats["complete"]:
        warnings.append(
            f"anchor {anchor} is PARTIAL ({anchor_stats['cells']}/{EXPECTED_CELLS}) — every published "
            f"comparison rests on it, so this must be investigated before the numbers are trusted")

    # Age of the data relative to the day the report runs. The scrape fires at
    # 02:00 UTC on the 1st and the report at 07:00 UTC the same day, so anything
    # older than a couple of days means the scrape did not land and the post is
    # about to re-publish last month's prices as if they were this month's.
    age_days = (report_date - latest).days
    if age_days > 2:
        warnings.append(
            f"the newest snapshot is {age_days} days old ({latest}) — this month's scrape "
            f"appears not to have landed")

    moved, compared = diff_cells(data, latest, anchor)
    prior = prior_complete(data, dates, latest)
    matrices = build_matrices(data, dates, latest, anchor)

    q = matrices.get("quarter") or {}
    if q.get("ref") and not q.get("on_target"):
        warnings.append(
            f"the quarter table compares against {q['ref']}, {q['actual_days_back']} days back "
            f"({q['off_target_days']} days off the 90-day target) — the nearest complete snapshot "
            f"to a true quarter, which the 2026-03-16 to 2026-06-30 scrape outage removed")
    elif not q.get("ref"):
        warnings.append("no earlier complete snapshot exists, so the quarter table cannot be built")

    return {
        "report_date": report_date.isoformat(),
        "anchor": anchor_stats,
        "latest": {**latest_stats, "age_days": age_days},
        "prior_complete": snapshot_stats(data, prior) if prior else None,
        "grid": {"municipalities": len(MUNI), "price_points": len(PRICE_POINTS),
                 "products": len(ALL_TIERS), "expected_cells": EXPECTED_CELLS},
        "price_basis": "net",   # ex-25% moms, as webPricingCalculator returns it
        "compared_cells": compared,
        "moved_count": len(moved),
        "moved_cells": moved,
        "matrices": matrices,
        "products": product_summary(moved, data, latest, anchor),
        "snapshots_total": len(dates),
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Excel: all data points
# ---------------------------------------------------------------------------
BOLD = Font(bold=True)
CELL_HEADERS = ["County", "Municipality", "Product", "Price point"]

# Diverging scale for % change: blue = cheaper, white = unchanged, red = dearer.
# The stops are FIXED at +/-10% rather than percentile-based, deliberately. A
# percentile scale re-normalises every time the workbook is rebuilt, so the same
# colour means a different thing each month and two sheets cannot be compared by
# eye. Fixed stops mean white always means "did not move" — which, on a grid where
# most cells are unchanged most months, is the whole point. Bigger moves (MAX at
# -20%) simply saturate, which is the correct emphasis.
PCT_SCALE = ColorScaleRule(
    start_type="num", start_value=-0.10, start_color="5B8FF9",   # blue
    mid_type="num", mid_value=0, mid_color="FFFFFF",
    end_type="num", end_value=0.10, end_color="F8696B")          # red


def _cell_rows(tiers=None):
    """The (muni, product, price) rows, grouped by county then municipality.

    `tiers` selects the product subset: CORE_TIERS for the listing packages,
    ADDON_TIERS for the upsells. Default is the full 420-cell grid.
    """
    for mid, (name, county) in sorted(MUNI.items(), key=lambda kv: (kv[1][1], kv[1][0])):
        for tier in (tiers or ALL_TIERS):
            for price in PRICE_POINTS:
                yield mid, name, county, tier, price


def _get(data, d, mid, tier, price):
    if d is None:
        return None
    return data.get(d, {}).get(mid, {}).get(tier, {}).get(price)


def _finish_sheet(ws, n_fixed, n_value_cols, number_format=None, rule=None):
    """Header styling, freeze panes, widths and conditional formatting."""
    for c in ws[3]:   # row 1 = subtitle, row 2 = spacer, row 3 = header
        c.font = BOLD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=4, column=n_fixed + 1).coordinate
    for col in range(1, n_fixed + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 17
    for col in range(n_fixed + 1, n_fixed + n_value_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    if not n_value_cols or ws.max_row < 4:
        return
    first = openpyxl.utils.get_column_letter(n_fixed + 1)
    last = openpyxl.utils.get_column_letter(n_fixed + n_value_cols)
    rng = f"{first}4:{last}{ws.max_row}"
    if number_format:
        for row in ws.iter_rows(min_row=4, min_col=n_fixed + 1, max_col=n_fixed + n_value_cols):
            for c in row:
                c.number_format = number_format
    if rule:
        ws.conditional_formatting.add(rng, rule)


def _pop_header(d, ref):
    """"2026-08-17 / vs 2026-07-12 (36d)".

    The gap is spelled out because the columns are NOT evenly spaced: the scrape ran
    weekly until 2026-03-16, then nothing until 2026-07-01, and is monthly from
    2026-08. Without the baseline in the header, the 2026-07-12 column looks like
    every other one while actually spanning 118 days across the outage.
    """
    return f"{d.isoformat()}\nvs {ref.isoformat()} ({(d - ref).days}d)"


def _pct_cell_sheet(wb, title, subtitle, data, cols, show_ref=True, tiers=None):
    """One row per grid cell, one column per (date, reference) pair, % change.

    `cols` is a list of (column_date, reference_date). A cell is left EMPTY when it
    is absent from either side — never 0 — so "not scraped" can always be told apart
    from "did not move", which is the distinction the whole sheet exists to show.
    """
    ws = wb.create_sheet(title)
    ws.append([subtitle])
    ws["A1"].font = BOLD
    ws.append([])
    ws.append(CELL_HEADERS + [(_pop_header(d, ref) if show_ref else d.isoformat())
                              for d, ref in cols])
    for mid, name, county, tier, price in _cell_rows(tiers):
        row = [county, name, tier, price]
        for d, ref in cols:
            cur = _get(data, d, mid, tier, price)
            base = _get(data, ref, mid, tier, price)
            row.append(None if (cur is None or base is None or not base)
                       else float(cur) / float(base) - 1.0)
        ws.append(row)
    _finish_sheet(ws, len(CELL_HEADERS), len(cols), "0.0%", PCT_SCALE)


def _prices_sheet(wb, data, dates, complete_set):
    """Raw prices, every snapshot including the partial ones (flagged in the header)."""
    ws = wb.create_sheet("Prices")
    ws.append(["Raw ad price in SEK, net of 25% moms, pay-when-removed basis. "
               "A header marked PARTIAL is a run that died mid-grid — its blanks are "
               "cells that were never scraped, not price removals."])
    ws["A1"].font = BOLD
    ws.append([])
    ws.append(CELL_HEADERS + [d.isoformat() + ("" if d in complete_set else " PARTIAL")
                              for d in dates])
    for mid, name, county, tier, price in _cell_rows():
        ws.append([county, name, tier, price]
                  + [_get(data, d, mid, tier, price) for d in dates])
    _finish_sheet(ws, len(CELL_HEADERS), len(dates), "#,##0")


def _county_pop_sheet(wb, data, complete):
    """County x product basket change, period on period — the audit trail for the
    two tables in the Slack post, computed by the SAME county_matrix function."""
    ws = wb.create_sheet("County PoP")
    ws.append(["Equal-weighted basket (each county's municipalities x all six price "
               "points), % change against the PREVIOUS COMPLETE snapshot. These are the "
               "same figures the Slack post's tables quote."])
    ws["A1"].font = BOLD
    ws.append([])
    pairs = list(zip(complete[:-1], complete[1:]))
    mats = {cur: county_matrix(data, cur, prev) for prev, cur in pairs}
    ws.append(["County", "Product"] + [_pop_header(cur, prev) for prev, cur in pairs])
    for county in COUNTIES:
        for tier in CORE_TIERS:
            row = [county, tier]
            for _, cur in pairs:
                r = next(x for x in mats[cur]["rows"] if x["county"] == county)
                v = r["tiers"][tier]
                row.append(None if v is None else v / 100.0)
            ws.append(row)
    for tier in CORE_TIERS:
        row = ["TOTAL (all counties)", tier]
        for _, cur in pairs:
            v = mats[cur]["total_row"][tier]
            row.append(None if v is None else v / 100.0)
        ws.append(row)
    _finish_sheet(ws, 2, len(pairs), "0.0%", PCT_SCALE)


def _snapshots_sheet(wb, data, dates, complete_set):
    """The index: what exists, what is complete, and where the gaps are."""
    ws = wb.create_sheet("Snapshots")
    ws.append(["Every snapshot in hemnet_adcostv2. Only COMPLETE runs are used as a "
               "baseline anywhere in this workbook or in the Slack post."])
    ws["A1"].font = BOLD
    ws.append([])
    ws.append(["Snapshot", "Cells", "Expected", "Municipalities", "Complete?",
               "Days since previous", "Used in PoP"])
    prev = None
    for d in dates:
        st = snapshot_stats(data, d)
        ws.append([d.isoformat(), st["cells"], EXPECTED_CELLS, st["munis"],
                   "yes" if st["complete"] else "NO",
                   (d - prev).days if prev else None,
                   "yes" if d in complete_set else "no"])
        prev = d
    _finish_sheet(ws, 7, 0)
    ws.freeze_panes = "A4"


def write_excel(data, dates, path, anchor):
    """Seven sheets, built for spotting CHANGE rather than reading levels.

    The previous single sheet colour-scaled RAW PRICES from min to max, which made a
    20,000 kr MAX cell permanently red and a 900 kr TOPLISTING permanently green — it
    encoded the price list, not its movement, so period-on-period change was invisible.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    complete = [d for d in dates if is_complete(data, d)]
    complete_set = set(complete)

    # Period on period: each complete snapshot against the previous COMPLETE one.
    # Partial runs are excluded as baselines for the same reason they are everywhere
    # else here — comparing against 2026-08-09 (35 of 420 cells) would paint ~385
    # never-scraped cells as changes.
    #
    # Only the COLUMN is required to be 2026+; its baseline is still the previous
    # complete snapshot whatever its date, so the first 2026 column keeps a real
    # comparison instead of rendering blank. Every header names its own baseline.
    pop_cols = [(d, ref) for d, ref in zip(complete[1:], complete[:-1])
                if d >= WORKBOOK_FROM]
    anchor_cols = [(d, anchor) for d in complete if d >= WORKBOOK_FROM]

    scale_note = ("White = unchanged, red = dearer, blue = cheaper (scale fixed at "
                  "+/-10% so colour means the same thing in every column and every "
                  "month). Blank = not scraped on one side. Columns start "
                  f"{WORKBOOK_FROM.isoformat()}; 2025 is on the Prices sheet.")

    _pct_cell_sheet(
        wb, "PoP change",
        "CORE LISTING FEES (Basic/Plus/Premium/Max) — % change vs the PREVIOUS "
        f"COMPLETE snapshot, per cell. {scale_note}",
        data, pop_cols, tiers=CORE_TIERS)

    _pct_cell_sheet(
        wb, "vs Anchor",
        f"CORE LISTING FEES — % change vs the fixed {anchor} baseline, per cell: the "
        f"cumulative index. Same colour scale as the PoP sheet. Columns start "
        f"{WORKBOOK_FROM.isoformat()}; the {anchor} baseline itself is NOT filtered.",
        data, anchor_cols, show_ref=False, tiers=CORE_TIERS)

    _pct_cell_sheet(
        wb, "Add-ons PoP",
        "UPSELLS (Paid republish / Toplisting / Toplisting 5 days) — bought ON TOP of "
        f"a listing package, never instead of one. {scale_note}",
        data, pop_cols, tiers=ADDON_TIERS)

    _pct_cell_sheet(
        wb, "Add-ons vs Anchor",
        f"UPSELLS — % change vs the fixed {anchor} baseline, per cell. Kept apart from "
        f"the core fees because the two moved in OPPOSITE directions: Max was cut ~21% "
        f"while the toplisting upsells rose ~20%.",
        data, anchor_cols, show_ref=False, tiers=ADDON_TIERS)

    _county_pop_sheet(wb, data, complete)
    _prices_sheet(wb, data, dates, complete_set)
    _snapshots_sheet(wb, data, dates, complete_set)
    wb.save(path)


# ---------------------------------------------------------------------------
# HTML heat map + ARPL
# ---------------------------------------------------------------------------
def cell_color(p):
    if p is None:
        return "#f2f2f2", "#999", "n/a"
    # diverging: red = price up, blue = price down; intensity by magnitude (cap 10%)
    mag = min(abs(p) / 0.10, 1.0)
    if p >= 0:
        bg = f"rgb(255,{int(255 - 120 * mag)},{int(255 - 120 * mag)})"
    else:
        bg = f"rgb({int(255 - 120 * mag)},{int(255 - 120 * mag)},255)"
    return bg, "#111", f"{p*100:+.1f}%"


def heat_table(title, subtitle, data_latest, data_ref, baseline):
    rows = []
    for county in COUNTIES:
        tds = [f"<th class='rowh'>{county}</th>"]
        for tier in CORE_TIERS:
            lp = weighted_price(data_latest, baseline, county, tier)
            rp = weighted_price(data_ref, baseline, county, tier) if data_ref else None
            bg, fg, txt = cell_color(pct(lp, rp))
            tip = f"{county} {tier}: {rp:.0f} → {lp:.0f} kr" if (lp and rp) else "n/a"
            tds.append(f"<td style='background:{bg};color:{fg}' title='{tip}'>{txt}</td>")
        rows.append("<tr>" + "".join(tds) + "</tr>")
    head = "<th></th>" + "".join(f"<th>{t.title()}</th>" for t in CORE_TIERS)
    return (f"<h2>{title}</h2><p class='sub'>{subtitle}</p>"
            f"<table class='heat'><tr>{head}</tr>{''.join(rows)}</table>")


def arpl_block(latest_date, data_latest, end2025_date, data_end, baseline):
    bl, pt_l = arpl(data_latest, baseline, CORE_TIERS)
    be, pt_e = arpl(data_end, baseline, CORE_TIERS) if data_end else (None, {})
    rows = []
    for tier in CORE_TIERS + ["BLENDED"]:
        if tier == "BLENDED":
            lv, ev = bl, be
        else:
            lv, ev = pt_l.get(tier), pt_e.get(tier)
        chg = pct(lv, ev)  # ratio — VAT cancels, so unaffected by MOMS
        _, _, ctxt = cell_color(chg)
        lg = (lv * MOMS) if lv else None   # gross (inc 25% moms)
        eg = (ev * MOMS) if ev else None
        rows.append(
            f"<tr><td class='rowh'>{tier.title()}</td>"
            f"<td><b>{('%.0f'%lg) if lg else 'n/a'}</b></td>"
            f"<td>{('%.0f'%eg) if eg else 'n/a'}</td>"
            f"<td>{ctxt}</td></tr>")
    return (f"<h2>Weighted ARPL (SEK / listing, inc. 25% moms)</h2>"
            f"<p class='sub'>Gross (VAT-inclusive), based on the last available listing "
            f"depth calculations in Feb-26. Weighted by this listing data (county × tier × price-band, "
            f"n={sum(sum(b.values()) for c in baseline.values() for b in c.values()):,} listings). "
            f"Blended = across Bas/Plus/Premium/Max.</p>"
            f"<table class='arpl'><tr><th>Tier</th><th>Latest ({latest_date})</th>"
            f"<th>End-2025 ({end2025_date})</th><th>Δ</th></tr>"
            f"{''.join(rows)}</table>")


def write_html(latest_date, prior_date, end2025_date, data, baseline, path, wow_ok):
    dl = data[latest_date]
    de = data.get(end2025_date)
    dp = data.get(prior_date)
    wow_sub = (f"latest {latest_date} vs previous complete snapshot {prior_date}"
               if wow_ok else
               f"n/a — no complete snapshot within {PRIOR_MAX_GAP_DAYS} days before "
               f"{latest_date}"
               + (f" (nearest complete run was {prior_date}, across a scrape gap)"
                  if prior_date else " (no earlier complete run at all)"))
    parts = [
        "<!doctype html><meta charset='utf-8'><title>Hemnet ad-cost heat map</title>",
        "<style>body{font-family:-apple-system,Segoe UI,Arial,sans-serif;margin:28px;color:#111}"
        "h1{margin:0 0 4px} .meta{color:#666;font-size:13px;margin-bottom:20px}"
        "h2{margin:26px 0 2px;font-size:17px} .sub{color:#777;font-size:12px;margin:0 0 8px}"
        "table{border-collapse:collapse;margin-bottom:8px} "
        ".heat td,.heat th{border:1px solid #e2e2e2;padding:7px 12px;text-align:center;font-size:13px;min-width:74px}"
        ".heat th{background:#fafafa} .rowh{background:#fafafa!important;text-align:left!important;font-weight:600}"
        ".arpl td,.arpl th{border:1px solid #e2e2e2;padding:6px 14px;text-align:right;font-size:13px}"
        ".arpl th{background:#fafafa} .legend{font-size:12px;color:#666;margin-top:6px}"
        "</style>",
        "<h1>Hemnet ad-cost — county heat map & ARPL</h1>",
        f"<div class='meta'>Basis: pay-when-removed price, net of 25% moms · pulled from "
        f"AdCostV2 · latest snapshot <b>{latest_date}</b> · anchor <b>{end2025_date}</b>. "
        f"Rows = the 8 priced counties; cols = ad package tier. "
        f"Cell = % change in the baseline-weighted price. "
        f"<span style='color:#c0392b'>red = cost up</span>, "
        f"<span style='color:#2c6fbf'>blue = cost down</span>.</div>",
        heat_table(f"% change vs the {end2025_date} anchor",
                   f"latest {latest_date} vs {end2025_date}", dl, de, baseline),
        heat_table("% change vs the previous complete snapshot", wow_sub,
                   dl, dp if wow_ok else None, baseline),
        arpl_block(latest_date, dl, end2025_date, de, baseline),
        "<p class='legend'>Gap 2026-03-16 → 2026-06-30 has no data (Hemnet prices are "
        "current-only; no backfill). Partial runs are never used as a baseline — a cell that "
        "was not scraped is not a price change.</p>",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Phase 28 ad-cost reporting")
    p.add_argument("--json", action="store_true",
                   help="print the machine-readable report on stdout (for adcost-report.js)")
    p.add_argument("--out-dir", default=None,
                   help=f"where to write the xlsx + heat map (default {OUT_DIR}; "
                        f"pass this with --json to get both from one DB pull)")
    p.add_argument("--report-date", default=None,
                   help="YYYY-MM-DD the report is being run for (default: today, UTC)")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    # In --json mode stdout is a data channel, so diagnostics go to stderr. Without
    # --json, --out-dir alone still writes artifacts and logs to stdout as before.
    write_files = (not args.json) or bool(args.out_dir)
    out_dir = args.out_dir or OUT_DIR
    log = (lambda *a: print(*a, file=sys.stderr)) if args.json else print

    report_date = (datetime.date.fromisoformat(args.report_date) if args.report_date
                   else datetime.datetime.now(datetime.timezone.utc).date())

    baseline = json.load(open(BASELINE, encoding="utf-8"))
    env = load_env()
    rows = fetch_rows(env)
    data, dates = build_cube(rows)
    if not dates:
        raise SystemExit("no AdCostV2 rows found")

    latest = dates[-1]
    prior = prior_complete(data, dates, latest)
    prior_ok = bool(prior and (latest - prior).days <= PRIOR_MAX_GAP_DAYS)

    if ANCHOR_DATE not in data:
        raise SystemExit(
            f"anchor {ANCHOR_DATE} has no rows — every published comparison is anchored on it, "
            f"so this is a hard stop rather than a silent fallback to another date")
    anchor = ANCHOR_DATE

    if write_files:
        os.makedirs(out_dir, exist_ok=True)
        xlsx = os.path.join(out_dir, "adcost-all-data.xlsx")
        html = os.path.join(out_dir, "adcost-heatmap.html")
        write_excel(data, dates, xlsx, anchor)
        write_html(latest, prior, anchor, data, baseline, html, prior_ok)
        log(f"wrote {xlsx}")
        log(f"wrote {html}")

    log(f"snapshots={len(dates)}  first={dates[0]}  latest={latest} "
        f"({cell_count(data[latest])}/{EXPECTED_CELLS} cells)")
    log(f"anchor={anchor}  prior_complete={prior}  prior_panel_valid={prior_ok}")

    if args.json:
        report = build_json(data, dates, latest, anchor, report_date)
        for w in report["warnings"]:
            log(f"WARNING: {w}")
        log(f"moved={report['moved_count']} of {report['compared_cells']} compared cells")
        # ensure_ascii=True deliberately: municipality names carry å/ä/ö, and this
        # stdout is a pipe read by Node. On Windows that pipe defaults to cp1252,
        # which mangles them; \uXXXX escapes decode identically on every platform.
        json.dump(report, sys.stdout, ensure_ascii=True)
        sys.stdout.write("\n")


if __name__ == "__main__":
    main()
